import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import asyncio
import sys
import urllib.parse
import os
import tempfile
import requests
from playwright.async_api import async_playwright
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

def scrape_seoul_auction(driver):
    print("\n--- 서울옥션 경매 정보 수집 시작 ---")
    url = "https://www.seoulauction.com/auction-list/upcoming"
    driver.get(url)
    
    try:
        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.CLASS_NAME, "auction_info"))
        )
    except Exception as e:
        print("서울옥션 데이터를 찾을 수 없거나 로딩 시간이 초과되었습니다.", e)
        return []

    time.sleep(2)
    auction_infos = driver.find_elements(By.CLASS_NAME, "auction_info")
    print(f"총 {len(auction_infos)}개의 서울옥션 경매 정보를 찾았습니다.")
    
    data_list = []
    for info in auction_infos:
        item = {}
        try:
            type_elements = info.find_elements(By.CLASS_NAME, "type")
            item['유형/상태'] = ", ".join([t.text for t in type_elements if t.text.strip()])
        except:
            item['유형/상태'] = ""
            
        try:
            title_element = info.find_element(By.CLASS_NAME, "title")
            item['경매명'] = title_element.text.strip()
        except:
            item['경매명'] = ""
            
        try:
            desc_area = info.find_element(By.CLASS_NAME, "description")
            dls = desc_area.find_elements(By.TAG_NAME, "dl")
            for dl in dls:
                dt_text = dl.find_element(By.TAG_NAME, "dt").text.strip()
                dd_text = dl.find_element(By.TAG_NAME, "dd").text.strip()
                item[dt_text] = dd_text
        except:
            pass
            
        item['바로가기 URL'] = f'=HYPERLINK("{url}", "{url}")'
        data_list.append(item)
        print(f"서울옥션 추출 완료: {item.get('경매명', '이름 없음')}")
        
    return data_list

def scrape_k_auction(driver):
    print("\n--- 케이옥션 출품작 정보 수집 시작 ---")
    url = "https://www.k-auction.com/Auction/Premium/225?page_size=10&page_type=P&auc_kind=2&auc_num=225&work_type=2672"
    driver.get(url)
    
    try:
        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.CLASS_NAME, "artwork"))
        )
    except Exception as e:
        print("케이옥션 데이터를 찾을 수 없거나 로딩 시간이 초과되었습니다.", e)
        return []

    # 경매 상단 정보 가져오기 (경매명, 일정, 장소)
    auction_title = ""
    auction_schedule = ""
    auction_location = ""
    
    try:
        subtop = driver.find_element(By.CLASS_NAME, "subtop-desc")
        auction_title = subtop.find_element(By.TAG_NAME, "h1").text.strip()
        
        p_tag = subtop.find_element(By.TAG_NAME, "p")
        auction_schedule = p_tag.find_element(By.TAG_NAME, "strong").text.strip()
        auction_location = p_tag.find_element(By.TAG_NAME, "span").text.strip()
        print(f"케이옥션 정보 파싱 완료: {auction_title}")
    except Exception as e:
        print("케이옥션 상단 정보를 가져오는 중 에러 발생:", e)

    # 전체 페이지 로딩을 위해 스크롤 다운 추가
    last_height = driver.execute_script("return document.body.scrollHeight")
    while True:
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(2)
        new_height = driver.execute_script("return document.body.scrollHeight")
        if new_height == last_height:
            break
        last_height = new_height

    time.sleep(2)
    artworks = driver.find_elements(By.CLASS_NAME, "artwork")
    print(f"총 {len(artworks)}개의 케이옥션 출품작 정보를 찾았습니다.")
    
    data_list = []
    for art in artworks:
        item = {}
        
        # 1. Lot 번호
        try:
            item['Lot'] = art.find_element(By.CLASS_NAME, "lot").text.strip()
        except:
            item['Lot'] = ""
            
        # 2. 이미지 URL
        try:
            img_tag = art.find_element(By.TAG_NAME, "img")
            img_url = img_tag.get_attribute("src")
            if not img_url or "sack_work_end.png" in img_url:
                img_url = img_tag.get_attribute("data-src")
            item['이미지 URL'] = f'=_xlfn.IMAGE("{img_url}")' if img_url else ""
        except:
            item['이미지 URL'] = ""

        # 3. 작가명
        try:
            item['작가명'] = art.find_element(By.CLASS_NAME, "card-title").text.strip()
        except:
            item['작가명'] = ""

        # 4. 작품명
        try:
            item['작품명'] = art.find_element(By.CLASS_NAME, "card-subtitle").text.strip()
        except:
            item['작품명'] = ""

        # 5. 재질 및 사이즈
        try:
            desc_ps = art.find_elements(By.CSS_SELECTOR, "p.description span")
            if len(desc_ps) >= 2:
                item['재질'] = desc_ps[0].text.strip()
                item['사이즈 및 연도'] = desc_ps[1].text.strip()
            elif len(desc_ps) == 1:
                item['재질/사이즈'] = desc_ps[0].text.strip()
        except:
            pass
            
        # 6. 추정가 및 시작가
        try:
            price_div = art.find_element(By.CLASS_NAME, "dotted")
            price_lis = price_div.find_elements(By.TAG_NAME, "li")
            
            # 텍스트들을 모아서 추정가와 시작가 찾기 (인덱스가 바뀔 수 있으므로 텍스트 기반 검색)
            for i, li in enumerate(price_lis):
                if "추정가" in li.text:
                    item['추정가 (KRW)'] = price_lis[i+1].text.strip()
                elif "시작가" in li.text:
                    item['시작가 (KRW)'] = price_lis[i+1].text.strip()
        except:
            pass

        # 7. 마감 시간
        try:
            # card-text 클래스들 중 마지막 것에 보통 마감 시간이 있음
            card_texts = art.find_elements(By.CLASS_NAME, "card-text")
            if len(card_texts) >= 2:
                item['마감 시간'] = card_texts[-1].text.strip()
        except:
            pass
            
        # 8. 디테일 페이지 URL
        try:
            link_tag = art.find_element(By.CLASS_NAME, "listimg")
            detail_url = link_tag.get_attribute("href")
            link = detail_url if detail_url else url
            item['바로가기 URL'] = f'=HYPERLINK("{link}", "{link}")'
        except:
            item['바로가기 URL'] = f'=HYPERLINK("{url}", "{url}")'
            
        # 9. 경매 상단 정보 일괄 추가
        item['경매명'] = auction_title
        item['일정'] = auction_schedule
        item['전시장소'] = auction_location

        data_list.append(item)
        print(f"케이옥션 추출 완료: Lot {item.get('Lot', '')} - {item.get('작가명', '')}")
        
    return data_list

def scrape_kan_auction(driver):
    print("\n--- 칸옥션 공지 정보 수집 시작 ---")
    url = "http://www.kanauction.kr/auction/going/main"
    driver.get(url)
    
    time.sleep(3) # 로딩 대기
    
    data_list = []
    item = {}
    
    try:
        # div 중 text-align:center 스타일을 가진 요소를 찾거나 텍스트 포함 요소 탐색
        elements = driver.find_elements(By.TAG_NAME, "div")
        info_text = ""
        for el in elements:
            if el.get_attribute("style") == "text-align:center" or el.get_attribute("style") == "text-align: center;":
                info_text = el.text.strip()
                if info_text and ("경매" in info_text or "칸옥션" in info_text):
                    break
        
        # 만약 style로 못 찾았다면 텍스트를 포함하는 div를 탐색 (가장 짧고 핵심이 되는 텍스트)
        if not info_text:
            for el in elements:
                text = el.text.strip()
                if "칸옥션 제" in text and "미술품경매" in text and "예정" in text:
                    info_text = text
                    break
                    
        item['공지 내용'] = info_text if info_text else "경매 정보를 찾을 수 없거나 아직 등록되지 않았습니다."
    except Exception as e:
        item['공지 내용'] = "에러 발생: " + str(e)
        
    item['바로가기 URL'] = f'=HYPERLINK("{url}", "{url}")'
    data_list.append(item)
    print("칸옥션 추출 완료.")
    
    return data_list

def scrape_myart_auction(driver):
    print("\n--- 마이아트옥션 공지 정보 수집 시작 ---")
    url = "https://myartauction.com/auctions/ongoing"
    driver.get(url)
    
    time.sleep(3) # 로딩 대기
    
    data_list = []
    item = {}
    
    try:
        page_source = driver.page_source
        if "NO CURRENT AUCTIONS" in page_source or "새로운 경매가 곧 시작됩니다" in page_source:
            item['공지 내용'] = "NO CURRENT AUCTIONS / 새로운 경매가 곧 시작됩니다"
        else:
            item['공지 내용'] = "현재 진행 중인 경매가 있습니다. (상세 내역은 홈페이지를 참고하세요)"
    except Exception as e:
        item['공지 내용'] = "에러 발생: " + str(e)
        
    item['바로가기 URL'] = f'=HYPERLINK("{url}", "{url}")'
    data_list.append(item)
    print("마이아트옥션 추출 완료.")
    
    return data_list

async def async_scrape_ebay(keyword):
    print(f"\n--- 이베이(eBay) 정보 수집 시작 (검색어: {keyword}) ---")
    nkw = urllib.parse.quote_plus(keyword)
    # 사용자가 제공한 더 상세한 URL 및 파라미터 적용
    url = f"https://www.ebay.com/sch/i.html?_nkw={nkw}&_sacat=0&_from=R40&_trksid=m570.l1313&_odkw={nkw}&_osacat=0"
    
    data_list = []
    async with async_playwright() as p:
        # 이베이 봇 탐지를 피하기 위해 브라우저 실행 (headless=True로 변경 가능)
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(
            user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36'
        )
        page = await context.new_page()
        
        try:
            await page.goto(url, wait_until="networkidle", timeout=30000)
            # 이미지 로딩을 위한 스크롤
            for _ in range(3):
                await page.mouse.wheel(0, 1000)
                await asyncio.sleep(1)
            
            # 아이템 컨테이너 찾기 (s-item 과 s-card 둘 다 지원)
            items = await page.query_selector_all("li.s-item, .s-card")
            for item in items:
                row = {}
                try:
                    # 제목 추출 (여러 셀렉터 시도)
                    title_el = await item.query_selector(".s-item__title, .s-card__title, div[role='heading']")
                    if not title_el: continue
                    title = await title_el.inner_text()
                    row['항목명'] = title.replace("새 창 또는 새 탭에서 열림", "").strip()
                    # 노이즈 제거
                    if not row['항목명'] or row['항목명'] in ["Shop on eBay", "eBay 상품 더보기", "관련 상품"]: continue
                except: continue
                
                try:
                    # 가격 추출
                    price_el = await item.query_selector(".s-item__price, .s-card__price, .s-item__price span")
                    row['가격 정보'] = await price_el.inner_text() if price_el else ""
                except: row['가격 정보'] = ""
                
                try:
                    # 배송 정보 추출 (동적 텍스트 검색 포함)
                    shipping_el = await item.query_selector(".s-item__shipping, .s-item__logisticsCost, .s-card__shipping, .s-item__free-shipping")
                    if not shipping_el:
                        # 텍스트 기반 검색으로 배송비 정보 시도
                        all_text = await item.inner_text()
                        if "배송" in all_text or "Shipping" in all_text:
                            shipping_el = await item.query_selector("span:has-text('배송'), span:has-text('Shipping')")
                            
                    row['배송 정보'] = await shipping_el.inner_text() if shipping_el else "Shipping info not found"
                except: row['배송 정보'] = "Shipping info not found"
                
                try:
                    # 링크 추출
                    link_el = await item.query_selector(".s-item__link, .s-card__link, a")
                    href = await link_el.get_attribute("href")
                    row['바로가기 URL'] = f'=HYPERLINK("{href}", "{href}")' if href else ""
                except: row['바로가기 URL'] = ""
                
                try:
                    # 이미지 추출 (폭넓은 이미지 태그 선택)
                    img_el = await item.query_selector(".s-item__image-img, .s-card__image-img, .s-card__link img, img")
                    img_url = await img_el.get_attribute("src") if img_el else ""
                    if not img_url or "placeholder" in img_url or "static" in img_url:
                        img_url = await img_el.get_attribute("data-src") if img_el else img_url
                    row['이미지 URL'] = img_url
                except: row['이미지 URL'] = ""
                
                data_list.append(row)
                print(f"이베이 추출 완료: {row['항목명'][:30]}...")
        except Exception as e:
            print(f"이베이 수집 중 에러 발생: {e}")
        finally:
            await browser.close()
            
    return data_list

def scrape_ebay_auction(keyword):
    # 비동기로 수집 수행 (윈도우 asyncio 이슈 해결)
    if sys.platform == 'win32':
        asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())
    return asyncio.run(async_scrape_ebay(keyword))

def save_to_excel_with_images(data_dict, filename):
    wb = Workbook()
    # 첫 번째 기본 시트 제거를 위해 나중에 처리
    first_sheet = True
    
    for sheet_name, data in data_dict.items():
        if first_sheet:
            ws = wb.active
            ws.title = sheet_name
            first_sheet = False
        else:
            ws = wb.create_sheet(title=sheet_name)
            
        df = pd.DataFrame(data)
        if df.empty:
            ws.append(["데이터가 없습니다."])
            continue
            
        # 헤더 추가
        ws.append(list(df.columns))
        
        # 데이터 및 이미지 추가
        for r_idx, row in enumerate(df.values, start=2):
            for c_idx, val in enumerate(row, start=1):
                col_name = df.columns[c_idx-1]
                
                # 이미지 처리 (사용자 요청에 따라 =IMAGE("URL") 함수 사용)
                # _xlfn.IMAGE 형식을 사용하여 @ 기호가 붙거나 오타가 나는 현상을 방지합니다.
                if col_name in ["이미지 URL", "이미지"]:
                    img_url = val
                    if img_url and isinstance(img_url, str) and img_url.startswith("http"):
                        # Excel IMAGE 함수 적용
                        ws.cell(row=r_idx, column=c_idx).value = f'=_xlfn.IMAGE("{img_url}")'
                        # 행 높이와 열 너비 조정 (이미지가 잘 보이도록)
                        ws.row_dimensions[r_idx].height = 80
                        ws.column_dimensions[get_column_letter(c_idx)].width = 25
                    else:
                        ws.cell(row=r_idx, column=c_idx, value=str(img_url) if img_url else "")
                
                # 하이퍼링크 처리
                elif "바로가기" in col_name and val and isinstance(val, str) and val.startswith("http"):
                    cell = ws.cell(row=r_idx, column=c_idx, value=val)
                    cell.hyperlink = val
                    cell.font = Font(color="0000FF", underline="single")
                
                else:
                    ws.cell(row=r_idx, column=c_idx, value=str(val) if val else "")
    
    wb.save(filename)
    print(f"파일 저장 완료: {filename}")

def main():
    print("브라우저를 초기화합니다...")
    options = webdriver.ChromeOptions()
    # 이베이 보안 우회를 위해 헤드리스 모드를 끄고 창을 생성합니다.
    # options.add_argument('--headless') 
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--disable-gpu')
    options.add_argument('--remote-debugging-port=9223')
    options.add_argument('--disable-extensions')
    options.add_argument('--disable-blink-features=AutomationControlled')
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)
    
    try:
        driver = webdriver.Chrome(options=options)
    except Exception as e:
        print(f"브라우저 초기화 실패: {e}")
        # 필요한 경우 다른 옵션으로 재시도 가능
        raise e
    driver.execute_cdp_cmd('Network.setUserAgentOverride', {
        "userAgent": 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36'
    })
    
    # 1. 사이트별 데이터 수집
    seoul_data = scrape_seoul_auction(driver)
    k_auction_data = scrape_k_auction(driver)
    kan_data = scrape_kan_auction(driver)
    myart_data = scrape_myart_auction(driver)
    
    driver.quit() # 기존 옥션 창 닫기
    
    # 이베이 데이터 수집
    ebay_keyword = "빈티지, 의약, 약국, 약, 약제상"
    ebay_data = scrape_ebay_auction(ebay_keyword)
    
    # 2. 엑셀 파일로 시트를 구분하여 저장
    if seoul_data or k_auction_data or kan_data or myart_data or ebay_data:
        excel_filename = "auction_upcoming_hyperlinks.xlsx"
        data_dict = {
            '서울옥션': seoul_data,
            '케이옥션': k_auction_data,
            '칸옥션': kan_data,
            '마이아트옥션': myart_data,
            '이베이(eBay)': ebay_data
        }
        
        save_to_excel_with_images(data_dict, excel_filename)
        print(f"\n>> 수집된 모든 데이터가 성공적으로 '{excel_filename}'에 탭으로 구분되어 저장되었습니다.")
    else:
        print("\n>> 수집된 데이터가 없어 엑셀 파일이 생성되지 않았습니다.")

if __name__ == "__main__":
    main()