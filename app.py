import streamlit as st
st.set_page_config(page_title="Auction Monitor", page_icon=r"C:\Users\kr2160068\.gemini\antigravity\brain\8a6cec09-c06a-49a3-9bb2-9f802b355b7b\media__1773296597207.png", layout="wide")
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import io
import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.worksheet.formula import ArrayFormula
import asyncio
import sys
import tempfile
import os
from playwright.async_api import async_playwright

# --- 브라우저 설정 ---
def init_driver():
    options = webdriver.ChromeOptions()
    # 이베이 등에서 봇 접근을 차단하는 것을 피하기 위해 헤드리스 모드를 끄고 창을 띄웁니다.
    # options.add_argument('--headless')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--disable-gpu')
    options.add_argument('--remote-debugging-port=9222')
    options.add_argument('--disable-extensions')
    # 자동화 탐지 우회 옵션
    options.add_argument('--disable-blink-features=AutomationControlled')
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)
    
    try:
        driver = webdriver.Chrome(options=options)
    except Exception as e:
        # 드라이버 생성 실패 시 상세 에러 출력
        print(f"Driver initialization failed: {e}")
        # 헤드리스 모드로 재시도 (환경에 따라 필요할 수 있음)
        # options.add_argument('--headless')
        # driver = webdriver.Chrome(options=options)
        raise e
    # user-agent 변경으로 봇 탐지율 낮춤
    driver.execute_cdp_cmd('Network.setUserAgentOverride', {
        "userAgent": 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36'
    })
    return driver

# --- 1. 서울옥션 데이터 수집 ---
def scrape_seoul(driver):
    url = "https://www.seoulauction.com/auction-list/upcoming"
    driver.get(url)
    
    try:
        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.CLASS_NAME, "auction_info"))
        )
    except Exception as e:
        return pd.DataFrame([{"에러": "데이터를 찾을 수 없습니다."}])

    time.sleep(2)
    auction_infos = driver.find_elements(By.CLASS_NAME, "auction_info")
    
    data_list = []
    for info in auction_infos:
        item = {}
        try:
            type_elements = info.find_elements(By.CLASS_NAME, "type")
            item['유형/상태'] = ", ".join([t.text for t in type_elements if t.text.strip()])
        except:
            item['유형/상태'] = ""
            
        try:
            title_element = info.find_element(By.CLASS_NAME, "title")
            item['경매명'] = title_element.text.strip()
        except:
            item['경매명'] = ""
            
        try:
            desc_area = info.find_element(By.CLASS_NAME, "description")
            dls = desc_area.find_elements(By.TAG_NAME, "dl")
            for dl in dls:
                dt_text = dl.find_element(By.TAG_NAME, "dt").text.strip()
                dd_text = dl.find_element(By.TAG_NAME, "dd").text.strip()
                item[dt_text] = dd_text
        except:
            pass
            
        item['선택'] = False
        item['바로가기 URL'] = url
        data_list.append(item)
        
    return pd.DataFrame(data_list)

# --- 3. 칸옥션 공지 정보 수집 ---
def scrape_kan(driver):
    url = "http://www.kanauction.kr/auction/going/main"
    driver.get(url)
    time.sleep(3)
    
    item = {}
    try:
        elements = driver.find_elements(By.TAG_NAME, "div")
        info_text = ""
        for el in elements:
            if el.get_attribute("style") in ["text-align:center", "text-align: center;"]:
                info_text = el.text.strip()
                if info_text and ("경매" in info_text or "칸옥션" in info_text):
                    break
        
        if not info_text:
            for el in elements:
                text = el.text.strip()
                if "칸옥션 제" in text and "미술품경매" in text and "예정" in text:
                    info_text = text
                    break
                    
        item['공지 내용'] = info_text if info_text else "경매 정보를 찾을 수 없거나 아직 등록되지 않았습니다."
    except Exception as e:
        item['공지 내용'] = "에러 발생"
        
    item['바로가기 URL'] = url
    return pd.DataFrame([item])

# --- 4. 마이아트옥션 공지 정보 수집 ---
def scrape_myart(driver):
    url = "https://myartauction.com/auctions/ongoing"
    driver.get(url)
    time.sleep(3)
    
    item = {}
    try:
        page_source = driver.page_source
        if "NO CURRENT AUCTIONS" in page_source or "새로운 경매가 곧 시작됩니다" in page_source:
            item['공지 내용'] = "NO CURRENT AUCTIONS / 새로운 경매가 곧 시작됩니다"
        else:
            item['공지 내용'] = "현재 진행 중인 경매가 있습니다. (상세 내역은 홈페이지를 참고하세요)"
    except Exception as e:
        item['공지 내용'] = "에러 발생"
        
    item['바로가기 URL'] = url
    return pd.DataFrame([item])

# --- 5. 이베이(eBay) 검색 결과 수집 ---
async def async_scrape_ebay(keyword):
    import urllib.parse
    # 콤마(,)가 포함된 검색어의 경우 이베이 URL에서 오류가 날 수 있으므로 공백으로 치환하고 정제합니다.
    clean_keyword = keyword.replace(",", " ").strip()
    while "  " in clean_keyword: clean_keyword = clean_keyword.replace("  ", " ")
    nkw = urllib.parse.quote_plus(clean_keyword)
    # 사용자가 제공한 더 상세한 URL 및 파라미터 적용
    url = f"https://www.ebay.com/sch/i.html?_nkw={nkw}&_sacat=0&_from=R40&_trksid=m570.l1313&_odkw={nkw}&_osacat=0"
    
    data_list = []
    async with async_playwright() as p:
        # 이베이 봇 탐지를 피하기 위해 브라우저 실행
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(
            user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36'
        )
        page = await context.new_page()
        
        try:
            await page.goto(url, wait_until="networkidle", timeout=30000)
            # 이미지 로딩을 위한 스크롤
            for _ in range(3):
                await page.mouse.wheel(0, 1000)
                await asyncio.sleep(1)
            
            # 아이템 컨테이너 찾기 (s-item 과 s-card 둘 다 지원)
            items = await page.query_selector_all("li.s-item, .s-card")
            for item in items:
                row = {'선택': False}
                try:
                    # 제목 추출 (여러 셀렉터 시도)
                    title_el = await item.query_selector(".s-item__title, .s-card__title, div[role='heading']")
                    if not title_el: continue
                    title = await title_el.inner_text()
                    row['항목명'] = title.replace("새 창 또는 새 탭에서 열림", "").strip()
                    # 노이즈 제거
                    if not row['항목명'] or row['항목명'] in ["Shop on eBay", "eBay 상품 더보기", "관련 상품"]: continue
                except: continue
                
                try:
                    # 가격 추출
                    price_el = await item.query_selector(".s-item__price, .s-card__price, .s-item__price span")
                    row['가격 정보'] = await price_el.inner_text() if price_el else ""
                except: row['가격 정보'] = ""
                
                try:
                    # 배송 정보 추출 (동적 텍스트 검색 포함)
                    shipping_el = await item.query_selector(".s-item__shipping, .s-item__logisticsCost, .s-card__shipping, .s-item__free-shipping")
                    if not shipping_el:
                        # 텍스트 기반 검색으로 배송비 정보 시도
                        all_text = await item.inner_text()
                        if "배송" in all_text or "Shipping" in all_text:
                            shipping_el = await item.query_selector("span:has-text('배송'), span:has-text('Shipping')")
                    
                    row['배송 정보'] = await shipping_el.inner_text() if shipping_el else "Shipping info not found"
                except: row['배송 정보'] = "Shipping info not found"
                
                try:
                    # 링크 추출
                    link_el = await item.query_selector(".s-item__link, .s-card__link, a")
                    href = await link_el.get_attribute("href")
                    row['바로가기'] = href if href else ""
                except: row['바로가기'] = ""
                
                try:
                    # 이미지 추출 (폭넓은 이미지 태그 선택)
                    img_el = await item.query_selector(".s-item__image-img, .s-card__image-img, .s-card__link img, img")
                    img_url = await img_el.get_attribute("src") if img_el else ""
                    if not img_url or "placeholder" in img_url or "static" in img_url:
                        img_url = await img_el.get_attribute("data-src") if img_el else img_url
                    row['이미지'] = img_url if img_url else ""
                except: row['이미지'] = ""
                
                data_list.append(row)
        except Exception as e:
            print(f"Ebay scraping error: {e}")
        finally:
            await browser.close()
            
    return data_list

def scrape_ebay(keyword):
    # 비동기로 수집 수행 (윈도우 asyncio 이슈 해결을 위해 ProactorEventLoopPolicy 사용)
    import sys
    if sys.platform == 'win32':
        asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())
        
    try:
        data_list = asyncio.run(async_scrape_ebay(keyword))
    except Exception as e:
        print(f"Asyncio run error: {e}")
        data_list = []
    
    empty_df = pd.DataFrame([{'항목명': '이베이 검색 결과가 없거나 수집에 실패했습니다.'}])
    if "선택" not in empty_df.columns:
        empty_df.insert(0, "선택", False)

    if not data_list:
        return empty_df
        
    df = pd.DataFrame(data_list)
    # 컬럼 순서 조정: 항목명, 이미지, 가격 정보, 배송 정보, 바로가기
    cols = ['항목명', '이미지', '가격 정보', '배송 정보', '바로가기']
    df = df[cols]
    df.insert(0, "선택", False) # 선택 컬럼 추가
    return df

# --- 공통 엑셀 다운로드 기능 ---
def add_df_to_sheet(ws, export_df, include_images=True):
    # 헤더 추가
    ws.append(list(export_df.columns))
    
    # 데이터 추가 루프
    for r_idx, row in enumerate(export_df.values, start=2):
        for c_idx, val in enumerate(row, start=1):
            col_name = export_df.columns[c_idx-1]
            
            # 이미지 처리 (사용자 요청에 따라 =IMAGE("URL") 함수 사용)
            # @기호가 붙는 현상을 방지하기 위해 ArrayFormula(전체 영역이 아닌 한 셀) 방식을 시도합니다.
            if include_images and col_name in ["이미지 URL", "이미지"]:
                img_url = val
                if img_url and isinstance(img_url, str) and img_url.startswith("http"):
                    # Excel IMAGE 함수 적용
                    # _xlfn.IMAGE 형식을 사용하면 @ 기호 없이 최신 함수를 인식시키는 데 도움이 됩니다.
                    ws.cell(row=r_idx, column=c_idx).value = f'=_xlfn.IMAGE("{img_url}")'
                    # 행 높이와 열 너비 조정
                    ws.row_dimensions[r_idx].height = 80
                    ws.column_dimensions[get_column_letter(c_idx)].width = 25
                else:
                    ws.cell(row=r_idx, column=c_idx, value=str(img_url) if img_url else "")
            
            # 하이퍼링크 처리 (바로가기 URL, 바로가기 등)
            elif "바로가기" in col_name and val and isinstance(val, str) and val.startswith("http"):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.hyperlink = val
                cell.font = Font(color="0000FF", underline="single")
            
            # 일반 텍스트
            else:
                ws.cell(row=r_idx, column=c_idx, value=str(val) if val else "")

def get_excel_data(df_to_download, sheet_name="목록"):
    output = io.BytesIO()
    export_df = df_to_download.copy()
    if "선택" in export_df.columns:
        export_df = export_df.drop(columns=["선택"])
    
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    add_df_to_sheet(ws, export_df)
    
    wb.save(output)
    return output.getvalue()

def get_multi_sheet_excel(dict_of_dfs):
    output = io.BytesIO()
    wb = Workbook()
    
    # 첫 번째 시트 삭제 (나중에 생성하기 위해)
    wb.remove(wb.active)
    
    for sheet_title, df in dict_of_dfs.items():
        if df is not None and not df.empty:
            ws = wb.create_sheet(title=sheet_title)
            export_df = df.copy()
            if "선택" in export_df.columns:
                export_df = export_df.drop(columns=["선택"])
            add_df_to_sheet(ws, export_df)
            
    wb.save(output)
    return output.getvalue()


# --- Streamlit UI 구현 ---

# 프리미엄 디자인을 위한 커스텀 CSS 적용
st.markdown("""
<style>
    /* 전체 배경 스타일 */
    .stApp {
        background: linear-gradient(135deg, #f8f9fa 0%, #e9ecef 100%);
    }
    
    /* 헤더 스타일 */
    .main-header {
        font-family: 'Playfair Display', serif;
        color: #1a1a1a;
        font-size: 3rem;
        font-weight: 700;
        margin-bottom: 0.5rem;
        text-align: center;
        border-bottom: 2px solid #d4af37;
        padding-bottom: 1rem;
    }
    
    /* 서브 타이틀 스타일 */
    .sub-header {
        color: #555;
        text-align: center;
        font-style: italic;
        margin-bottom: 2rem;
    }
    
    /* 스크롤바 스타일 커스텀 */
    ::-webkit-scrollbar {
        width: 8px;
        height: 8px;
    }
    ::-webkit-scrollbar-track {
        background: #f1f1f1;
        border-radius: 10px;
    }
    ::-webkit-scrollbar-thumb {
        background: #d4af37;
        border-radius: 10px;
    }
    ::-webkit-scrollbar-thumb:hover {
        background: #b8962d;
    }

    /* 전체 앱 컨테이너 스크롤 허용 */
    .main .block-container {
        overflow: auto !important;
    }

    /* 카드/컨테이너 스타일 (Glassmorphism) */
    .stDataFrame, .stTable {
        border-radius: 10px;
        box-shadow: 0 4px 6px rgba(0,0,0,0.05);
        background-color: rgba(255, 255, 255, 0.8);
    }
    
    /* 탭 스타일 개선 */
    .stTabs [data-baseweb="tab-list"] {
        gap: 20px;
        justify-content: center;
    }
    
    .stTabs [data-baseweb="tab"] {
        height: 50px;
        padding-top: 10px;
        padding-bottom: 10px;
        background-color: transparent;
        border-radius: 4px 4px 0px 0px;
        gap: 10px;
        font-weight: 600;
    }
    
    .stTabs [aria-selected="true"] {
        background-color: rgba(212, 175, 55, 0.1) !important;
        border-bottom: 3px solid #d4af37 !important;
    }

    /* 버튼 스타일 커스텀 */
    div.stButton > button:first-child {
        background-color: #1a1a1a;
        color: white;
        border-radius: 20px;
        padding: 0.5rem 2rem;
        border: none;
        transition: all 0.3s ease;
    }
    
    div.stButton > button:first-child:hover {
        background-color: #d4af37;
        color: #1a1a1a;
        transform: translateY(-2px);
    }

    /* 연관성 버튼 특수 스타일 */
    div.stButton > button.relevance-btn {
        background-color: #d4af37 !important;
        color: #1a1a1a !important;
        font-weight: bold;
    }
</style>
""", unsafe_allow_html=True)

# --- 콜백 함수 정의 (데이터 에디터 동기화용) ---
def sync_ebay_ko_editor():
    if "ebay_ko_editor" in st.session_state:
        edits = st.session_state["ebay_ko_editor"]["edited_rows"]
        for idx_str, change in edits.items():
            idx = int(idx_str)
            for col, val in change.items():
                st.session_state['df_ebay_ko'].at[idx, col] = val

def sync_ebay_en_editor():
    if "ebay_en_editor" in st.session_state:
        edits = st.session_state["ebay_en_editor"]["edited_rows"]
        for idx_str, change in edits.items():
            idx = int(idx_str)
            for col, val in change.items():
                st.session_state['df_ebay_en'].at[idx, col] = val

# 상단 헤더 및 홈 버튼 레이아웃
col_h1, col_h2 = st.columns([10, 2])
with col_h1:
    st.markdown('<h1 class="main-header" style="text-align: left; border-bottom: none;">Auction Monitor</h1>', unsafe_allow_html=True)
    st.markdown('<p class="sub-header" style="text-align: left;">Premium Art Auction Aggregator & Selection Tool</p>', unsafe_allow_html=True)
with col_h2:
    st.write("##") # 상단 여백
    if st.button("🏠 홈 (초기화)", use_container_width=True, help="데이터와 설정을 모두 초기화하고 홈으로 돌아갑니다."):
        st.session_state.clear()
        st.rerun()
st.markdown('<div style="border-bottom: 2px solid #d4af37; margin-bottom: 2rem;"></div>', unsafe_allow_html=True)

st.info("💡 **'데이터 수집 시작'** 버튼을 누르면 서울옥션, 케이옥션, 칸옥션, 마이아트옥션, 이베이의 실시간 정보를 가져옵니다.")

default_ebay_ko = '빈티지, 의약, 약국, 약, 약제상'
default_ebay_en = "apothecary, bottle, jar, tool, scale, medicine"

col_kw1, col_kw2 = st.columns(2)
with col_kw1:
    ebay_keyword_ko = st.text_input("🔍 이베이 한국어 검색어", value=default_ebay_ko)
with col_kw2:
    ebay_keyword_en = st.text_input("🔍 이베이 영어 검색어", value=default_ebay_en)

col_btn_1, col_btn_2, col_btn_3 = st.columns([1, 1, 1])
with col_btn_2:
    start_btn = st.button("🚀 실시간 데이터 수집", type="primary", use_container_width=True)

# 다이내믹 위젯들이 st.tabs의 인덱스를 흔들지 않도록 placeholder 예약
global_export_placeholder = st.empty()
status_placeholder = st.empty()
info_placeholder = st.empty()

if start_btn:
    # 엑셀 생성 상태 초기화
    st.session_state['ebay_ko_excel_ready'] = False
    st.session_state['ebay_en_excel_ready'] = False
    st.session_state['all_excel_ready'] = False
    
    driver = init_driver()
    
    # 세션 스테이트(Session State)를 활용하여 데이터 저장
    with status_placeholder.status("로봇이 열심히 데이터를 수집 중입니다...", expanded=True) as status:
        st.write("🏃 서울옥션 데이터 수집 중...")
        st.session_state['df_seoul'] = scrape_seoul(driver)
        
        st.write("🏃 칸옥션 공지 가져오는 중...")
        st.session_state['df_kan'] = scrape_kan(driver)
        
        st.write("🏃 마이아트옥션 공지 가져오는 중...")
        st.session_state['df_myart'] = scrape_myart(driver)
        driver.quit() # 기존 옥션들 창 닫기
        
        st.write(f"🏃 이베이 한국어 검색 중... (검색어: {ebay_keyword_ko})")
        st.session_state['df_ebay_ko'] = scrape_ebay(ebay_keyword_ko)
        
        st.write(f"🏃 이베이 영어 검색 중... (검색어: {ebay_keyword_en})")
        st.session_state['df_ebay_en'] = scrape_ebay(ebay_keyword_en)
        
        status.update(label="✅ 모든 데이터 수집이 완료되었습니다!", state="complete", expanded=False)
        
    # 이미 윗줄에서 닫혔으므로 생략

# 수집된 데이터가 있으면 화면 상단(버튼 하단)에 전체 다운로드 버튼 배치
if "df_seoul" in st.session_state:
    with global_export_placeholder.container():
        st.divider()
        col1, col2 = st.columns([1, 1])
        with col1:
            if st.button("📊 모든 옥션 데이터 통합 엑셀 생성", use_container_width=True):
                with st.spinner("이미지를 포함한 통합 엑셀을 생성 중 (시간이 소요됩니다)..."):
                    dfs = {
                        "서울옥션": st.session_state.get('df_seoul'),
                        "칸옥션": st.session_state.get('df_kan'),
                        "마이아트옥션": st.session_state.get('df_myart'),
                        "이베이(한국어)": st.session_state.get('df_ebay_ko'),
                        "이베이(영어)": st.session_state.get('df_ebay_en')
                    }
                    all_excel_bytes = get_multi_sheet_excel(dfs)
                    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                        tmp.write(all_excel_bytes)
                        st.session_state['all_excel_path'] = tmp.name
                    st.session_state['all_excel_ready'] = True
                    st.toast("✅ 통합 엑셀 생성이 완료되었습니다!")
        
        with col2:
            if st.session_state.get('all_excel_ready') and 'all_excel_path' in st.session_state:
                with open(st.session_state['all_excel_path'], "rb") as f:
                    st.download_button(
                        label="📥 통합 엑셀 파일 다운로드",
                        data=f.read(),
                        file_name="all_auctions_report.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        type="primary"
                    )
        st.divider()

    tab1, tab3, tab4, tab5, tab6 = st.tabs(["🏛️ 서울옥션", "🏛️ 칸옥션", "🏛️ 마이아트옥션", "🛒 이베이(한국어 검색)", "🛒 이베이(영어 검색)"])

    with tab1:
        st.subheader("서울옥션 예정 경매")
        st.dataframe(
            st.session_state['df_seoul'],
            column_config={
                "바로가기 URL": st.column_config.LinkColumn("해당 페이지로 이동")
            },
            use_container_width=True,
            hide_index=True
        )

    with tab3:
        st.subheader("칸옥션 공지사항")
        st.dataframe(
            st.session_state['df_kan'],
            column_config={
                "바로가기 URL": st.column_config.LinkColumn("홈페이지 열기")
            },
            use_container_width=True,
            hide_index=True
        )

    with tab4:
        st.subheader("마이아트옥션 진행경매 안내")
        st.dataframe(
            st.session_state['df_myart'],
            column_config={
                "바로가기 URL": st.column_config.LinkColumn("홈페이지 열기")
            },
            use_container_width=True,
            hide_index=True
        )

    with tab5:
        st.subheader("이베이(한국어 검색) 결과")
        if 'df_ebay_ko' in st.session_state and not st.session_state['df_ebay_ko'].empty:
            df_ebay_ko = st.session_state['df_ebay_ko']
            
            # 전체 선택 및 연관성 버튼
            col_select_ebay_ko, col_rel_ebay_ko, col_space_ebay_ko = st.columns([1.5, 1.5, 7])
            with col_select_ebay_ko:
                chk_all_ebay_ko = st.checkbox("전체 선택", key="btn_select_all_ebay_ko", help="모든 항목을 선택하거나 해제합니다.")
                if chk_all_ebay_ko != st.session_state.get('prev_chk_all_ebay_ko', False):
                    st.session_state['df_ebay_ko']['선택'] = chk_all_ebay_ko
                    st.session_state['prev_chk_all_ebay_ko'] = chk_all_ebay_ko

            with col_rel_ebay_ko:
                if st.button("✨ 연관성 높음", key="btn_rel_ebay_ko", help="약장, 약합, 약재 등 관련 항목을 자동 선택합니다."):
                    rel_keywords = ["약장", "약합", "약재", "십장생", "건강", "장수", "의원", "의서", "약학", "의학", "유의", "약"]
                    def check_rel_ebay(row):
                        text = f"{row.get('항목명', '')} {row.get('이미지', '')}"
                        return any(kw in text for kw in rel_keywords)
                    st.session_state['df_ebay_ko']['선택'] = st.session_state['df_ebay_ko'].apply(check_rel_ebay, axis=1)
                    st.toast("✅ 연관성 높은 항목들이 선택되었습니다.")

            # 데이터 에디터 출력
            edited_ebay_ko_df = st.data_editor(
                st.session_state['df_ebay_ko'],
                column_config={
                    "선택": st.column_config.CheckboxColumn("선택", default=False),
                    "바로가기": st.column_config.LinkColumn("상품 페이지 열기"),
                    "이미지": st.column_config.ImageColumn("이미지")
                },
                use_container_width=True,
                hide_index=True,
                height=600, # 스크롤바 확보
                key="ebay_ko_editor",
                on_change=sync_ebay_ko_editor
            )
            
            selected_ebay_ko_df = st.session_state['df_ebay_ko'][st.session_state['df_ebay_ko']["선택"] == True]
            df_for_ebay_ko_dl = selected_ebay_ko_df if not selected_ebay_ko_df.empty else edited_ebay_ko_df
            
            if st.button("📊 다운로드용 엑셀 파일 생성", key="btn_gen_ebay_ko"):
                with st.spinner("이미지를 포함한 이베이 엑셀 파일을 생성 중입니다..."):
                    excel_bytes = get_excel_data(df_for_ebay_ko_dl)
                    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                        tmp.write(excel_bytes)
                        st.session_state['ebay_ko_excel_path'] = tmp.name
                    st.session_state['ebay_ko_excel_ready'] = True
            
            if st.session_state.get('ebay_ko_excel_ready') and 'ebay_ko_excel_path' in st.session_state:
                with open(st.session_state['ebay_ko_excel_path'], "rb") as f:
                    ebay_ko_excel_data = f.read()
                ebay_ko_dl_label = "📥 선택한 항목 엑셀 다운로드" if not selected_ebay_ko_df.empty else "📥 전체 항목 엑셀 다운로드"
                st.download_button(
                    label=ebay_ko_dl_label,
                    data=ebay_ko_excel_data,
                    file_name="ebay_items_ko.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="ebay_ko_download",
                    type="primary"
                )
        else:
            st.warning("이베이 한국어 검색 결과가 없거나 수집에 실패했습니다.")

    with tab6:
        st.subheader("이베이(영어 검색) 결과")
        if 'df_ebay_en' in st.session_state and not st.session_state['df_ebay_en'].empty:
            df_ebay_en = st.session_state['df_ebay_en']
            
            # 전체 선택 버튼 (요청에 따라 연관성 버튼 제거)
            col_select_ebay_en, col_space_ebay_en = st.columns([1.5, 8.5])
            with col_select_ebay_en:
                chk_all_ebay_en = st.checkbox("전체 선택", key="btn_select_all_ebay_en", help="모든 항목을 선택하거나 해제합니다.")
                if chk_all_ebay_en != st.session_state.get('prev_chk_all_ebay_en', False):
                    st.session_state['df_ebay_en']['선택'] = chk_all_ebay_en
                    st.session_state['prev_chk_all_ebay_en'] = chk_all_ebay_en

            # 데이터 에디터 출력
            edited_ebay_en_df = st.data_editor(
                st.session_state['df_ebay_en'],
                column_config={
                    "선택": st.column_config.CheckboxColumn("선택", default=False),
                    "바로가기": st.column_config.LinkColumn("상품 페이지 열기"),
                    "이미지": st.column_config.ImageColumn("이미지")
                },
                use_container_width=True,
                hide_index=True,
                height=600, # 스크롤바 확보
                key="ebay_en_editor",
                on_change=sync_ebay_en_editor
            )
            
            selected_ebay_en_df = st.session_state['df_ebay_en'][st.session_state['df_ebay_en']["선택"] == True]
            df_for_ebay_en_dl = selected_ebay_en_df if not selected_ebay_en_df.empty else edited_ebay_en_df
            
            if st.button("📊 다운로드용 엑셀 파일 생성", key="btn_gen_ebay_en"):
                with st.spinner("이미지를 포함한 이베이 엑셀 파일을 생성 중입니다..."):
                    excel_bytes = get_excel_data(df_for_ebay_en_dl)
                    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                        tmp.write(excel_bytes)
                        st.session_state['ebay_en_excel_path'] = tmp.name
                    st.session_state['ebay_en_excel_ready'] = True
            
            if st.session_state.get('ebay_en_excel_ready') and 'ebay_en_excel_path' in st.session_state:
                with open(st.session_state['ebay_en_excel_path'], "rb") as f:
                    ebay_en_excel_data = f.read()
                ebay_en_dl_label = "📥 선택한 항목 엑셀 다운로드" if not selected_ebay_en_df.empty else "📥 전체 항목 엑셀 다운로드"
                st.download_button(
                    label=ebay_en_dl_label,
                    data=ebay_en_excel_data,
                    file_name="ebay_items_en.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="ebay_en_download",
                    type="primary"
                )
        else:
            st.warning("이베이 영어 검색 결과가 없거나 수집에 실패했습니다.")
