import streamlit as st
import os

# --- 클라우드 환경에서 Playwright 브라우저 설치 보장 ---
if not os.path.exists(os.path.expanduser("~/.cache/ms-playwright")):
    with st.spinner("최초 실행을 위해 브라우저 엔진을 설치 중입니다. 잠시만 기다려 주세요..."):
        os.system("playwright install chromium")

st.set_page_config(page_title="Auction Monitor", page_icon="🏢", layout="wide")
import pandas as pd
import time
import io
import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import asyncio
import sys
import tempfile
from playwright.async_api import async_playwright

# --- Windows 환경 Playwright asyncio.subprocess 에러 해결 (NotImplementedError 방지) ---
if sys.platform == "win32":
    asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())


async def apply_stealth(page):
    await page.add_init_script("""
        Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
        window.chrome = { runtime: {} };
        Object.defineProperty(navigator, 'languages', { get: () => ['ko-KR', 'ko', 'en-US', 'en'] });
        Object.defineProperty(navigator, 'plugins', { get: () => [1, 2, 3, 4, 5] });
    """)

# --- 브라우저 설정 helper ---
async def get_browser_context(p):
    browser = await p.chromium.launch(headless=True)
    context = await browser.new_context(
        user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36',
        viewport={'width': 1280, 'height': 800},
        extra_http_headers={
            "Accept-Language": "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8",
        }
    )
    return browser, context

# --- 1. 서울옥션 데이터 수집 ---
# --- 1. 서울옥션 데이터 수집 (Playwright) ---
async def async_scrape_seoul():
    url = "https://www.seoulauction.com/auction-list/upcoming"
    data_list = []
    async with async_playwright() as p:
        browser, context = await get_browser_context(p)
        page = await context.new_page()
        await apply_stealth(page)
        try:
            await page.goto(url, wait_until="networkidle", timeout=60000)
            await page.wait_for_selector(".auction_info", timeout=15000)
            
            auction_infos = await page.query_selector_all(".auction_info")
            for info in auction_infos:
                item = {}
                try:
                    types = await info.query_selector_all(".type")
                    type_texts = [await t.inner_text() for t in types]
                    item['유형/상태'] = ", ".join([t.strip() for t in type_texts if t.strip()])
                except: item['유형/상태'] = ""
                
                try:
                    title_el = await info.query_selector(".title")
                    item['경매명'] = (await title_el.inner_text()).strip()
                except: item['경매명'] = ""
                
                try:
                    dls = await info.query_selector_all(".description dl")
                    for dl in dls:
                        dt = await dl.query_selector("dt")
                        dd = await dl.query_selector("dd")
                        dt_text = (await dt.inner_text()).strip()
                        dd_text = (await dd.inner_text()).strip()
                        item[dt_text] = dd_text
                except: pass
                
                item['선택'] = False
                item['바로가기 URL'] = url
                data_list.append(item)
        except Exception as e:
            print(f"Seoul Auction error: {e}")
        finally:
            await browser.close()
    return pd.DataFrame(data_list)



# --- 3. 칸옥션 공지 정보 수집 ---
# --- 3. 칸옥션 공지 (Playwright) ---
async def async_scrape_kan():
    url = "http://www.kanauction.kr/auction/going/main"
    item = {'바로가기 URL': url}
    async with async_playwright() as p:
        browser, context = await get_browser_context(p)
        page = await context.new_page()
        await apply_stealth(page)
        try:
            await page.goto(url, wait_until="networkidle", timeout=30000)
            divs = await page.query_selector_all("div")
            info_text = ""
            for div in divs:
                txt = (await div.inner_text()).strip()
                if "칸옥션" in txt and "경매" in txt:
                    info_text = txt
                    break
            item['공지 내용'] = info_text if info_text else "경매 정보를 아직 등록되지 않았습니다."
        except:
            item['공지 내용'] = "정보를 불러올 수 없습니다."
        finally:
            await browser.close()
    return pd.DataFrame([item])

# --- 4. 마이아트옥션 공지 (Playwright) ---
async def async_scrape_myart():
    url = "https://myartauction.com/auctions/ongoing"
    item = {'바로가기 URL': url}
    async with async_playwright() as p:
        browser, context = await get_browser_context(p)
        page = await context.new_page()
        await apply_stealth(page)
        try:
            await page.goto(url, wait_until="networkidle", timeout=30000)
            source = await page.content()
            if "NO CURRENT AUCTIONS" in source or "새로운 경매가 곧 시작됩니다" in source:
                item['공지 내용'] = "NO CURRENT AUCTIONS / 새로운 경매가 곧 시작됩니다"
            else:
                item['공지 내용'] = "현재 진행 중인 경매가 있습니다."
        except:
            item['공지 내용'] = "에러 발생"
        finally:
            await browser.close()
    return pd.DataFrame([item])

# --- 5. 이베이(eBay) 검색 결과 수집 ---
async def async_scrape_ebay(keyword):
    import urllib.parse
    # 콤마(,) 정제 (제거하지 않고 URL 인코딩에 그대로 반영되도록 수정)
    clean_keyword = keyword.strip()
    while "  " in clean_keyword: clean_keyword = clean_keyword.replace("  ", " ")
    nkw = urllib.parse.quote_plus(clean_keyword)
    url = f"https://www.ebay.com/sch/i.html?_nkw={nkw}&_sacat=0&_from=R40&_trksid=m570.l1313&_odkw={nkw}&_osacat=0"
    
    data_list = []
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(
            user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36',
            viewport={'width': 1280, 'height': 800},
            extra_http_headers={
                "Accept-Language": "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7",
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8",
            }
        )
        page = await context.new_page()
        # 스텔스 모드 적용
        await apply_stealth(page)
        
        try:
            # 타임아웃을 늘리고 대기 전략을 변경합니다.
            await page.goto(url, wait_until="domcontentloaded", timeout=60000)
            await asyncio.sleep(3) 
            
            # 스크롤링
            for _ in range(3):
                await page.evaluate("window.scrollBy(0, 800)")
                await asyncio.sleep(1)
            
            # 항목 대기
            try:
                await page.wait_for_selector("li.s-item, li.s-card", timeout=10000)
            except: pass

            items = await page.query_selector_all("li.s-item, li.s-card")
            for item in items:
                row = {'선택': False}
                try:
                    title_el = await item.query_selector(".s-item__title, .s-card__title, div[role='heading']")
                    if not title_el: continue
                    title = await title_el.inner_text()
                    row['항목명'] = title.replace("새 창 또는 새 탭에서 열림", "").strip()
                    if not row['항목명'] or row['항목명'] in ["Shop on eBay", "eBay 상품 더보기", "관련 상품"]: continue
                except: continue
                
                try:
                    price_el = await item.query_selector(".s-item__price, .s-card__price, .s-item__price span")
                    row['가격 정보'] = await price_el.inner_text() if price_el else ""
                except: row['가격 정보'] = ""
                
                try:
                    shipping_el = await item.query_selector(".s-item__shipping, .s-item__logisticsCost, .s-card__shipping, .s-item__free-shipping")
                    if not shipping_el:
                        all_text = await item.inner_text()
                        if "배송" in all_text or "Shipping" in all_text:
                            shipping_el = await item.query_selector("span:has-text('배송'), span:has-text('Shipping')")
                    
                    row['배송 정보'] = await shipping_el.inner_text() if shipping_el else "Shipping info not found"
                except: row['배송 정보'] = "Shipping info not found"
                
                try:
                    link_el = await item.query_selector(".s-item__link, .s-card__link, a")
                    href = await link_el.get_attribute("href")
                    row['바로가기'] = href if href else ""
                except: row['바로가기'] = ""
                
                try:
                    img_el = await item.query_selector(".s-item__image-img, .s-card__image-img, .s-card__link img, img")
                    img_url = await img_el.get_attribute("src") if img_el else ""
                    if not img_url or "placeholder" in img_url or "static" in img_url:
                        img_url = await img_el.get_attribute("data-src") if img_el else img_url
                    row['이미지'] = img_url if img_url else ""
                except: row['이미지'] = ""
                
                data_list.append(row)
        except Exception as e:
            print(f"Ebay scraping error: {e}")
        finally:
            await browser.close()
            
    return data_list

def scrape_ebay(keyword):
    try:
        data_list = asyncio.run(async_scrape_ebay(keyword))
    except Exception as e:
        print(f"Asyncio run error: {e}")
        data_list = []
    
    empty_df = pd.DataFrame([{'항목명': '이베이 검색 결과가 없거나 수집에 실패했습니다.'}])
    if "선택" not in empty_df.columns:
        empty_df.insert(0, "선택", False)

    if not data_list:
        return empty_df
        
    df = pd.DataFrame(data_list)
    cols = ['항목명', '이미지', '가격 정보', '배송 정보', '바로가기']
    df = df[cols]
    df.insert(0, "선택", False)
    return df

# --- 공통 엑셀 다운로드 기능 ---
def add_df_to_sheet(ws, export_df, include_images=True):
    ws.append(list(export_df.columns))
    for r_idx, row in enumerate(export_df.values, start=2):
        for c_idx, val in enumerate(row, start=1):
            col_name = export_df.columns[c_idx-1]
            if include_images and col_name in ["이미지 URL", "이미지"]:
                img_url = val
                if img_url and isinstance(img_url, str) and img_url.startswith("http"):
                    ws.cell(row=r_idx, column=c_idx).value = f'=_xlfn.IMAGE("{img_url}")'
                    ws.row_dimensions[r_idx].height = 80
                    ws.column_dimensions[get_column_letter(c_idx)].width = 25
                else:
                    ws.cell(row=r_idx, column=c_idx, value=str(img_url) if img_url else "")
            elif "바로가기" in col_name and val and isinstance(val, str) and val.startswith("http"):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.hyperlink = val
                cell.font = Font(color="0000FF", underline="single")
            else:
                ws.cell(row=r_idx, column=c_idx, value=str(val) if val else "")

def get_excel_data(df_to_download, sheet_name="목록"):
    output = io.BytesIO()
    export_df = df_to_download.copy()
    if "선택" in export_df.columns:
        export_df = export_df.drop(columns=["선택"])
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    add_df_to_sheet(ws, export_df)
    wb.save(output)
    return output.getvalue()

def get_multi_sheet_excel(dict_of_dfs):
    output = io.BytesIO()
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_title, df in dict_of_dfs.items():
        if df is not None and not df.empty:
            ws = wb.create_sheet(title=sheet_title)
            export_df = df.copy()
            if "선택" in export_df.columns:
                export_df = export_df.drop(columns=["선택"])
            add_df_to_sheet(ws, export_df)
    wb.save(output)
    return output.getvalue()


# --- Streamlit UI ---
st.markdown("""
<style>
    .stApp { background: linear-gradient(135deg, #f8f9fa 0%, #e9ecef 100%); }
    .main-header { font-family: 'Playfair Display', serif; color: #1a1a1a; font-size: 3rem; font-weight: 700; text-align: center; border-bottom: 2px solid #d4af37; padding-bottom: 1rem; }
    .sub-header { color: #555; text-align: center; font-style: italic; margin-bottom: 2rem; }
    ::-webkit-scrollbar { width: 8px; height: 8px; }
    ::-webkit-scrollbar-track { background: #f1f1f1; border-radius: 10px; }
    ::-webkit-scrollbar-thumb { background: #d4af37; border-radius: 10px; }
    .main .block-container { overflow: auto !important; }
    .stDataFrame, .stTable { border-radius: 10px; box-shadow: 0 4px 6px rgba(0,0,0,0.05); background-color: rgba(255, 255, 255, 0.8); }
    .stTabs [data-baseweb="tab-list"] { gap: 20px; justify-content: center; }
    .stTabs [aria-selected="true"] { background-color: rgba(212, 175, 55, 0.1) !important; border-bottom: 3px solid #d4af37 !important; }
    div.stButton > button:first-child { background-color: white; color: #1a1a1a; border: 1px solid #1a1a1a; border-radius: 20px; padding: 0.5rem 2rem; transition: all 0.3s ease; }
    div.stButton > button:first-child:hover { background-color: #f8f9fa; color: #d4af37; border-color: #d4af37; transform: translateY(-2px); box-shadow: 0 4px 10px rgba(0,0,0,0.05); }
</style>
""", unsafe_allow_html=True)

def sync_ebay_ko_editor():
    if "ebay_ko_editor" in st.session_state:
        edits = st.session_state["ebay_ko_editor"]["edited_rows"]
        for idx_str, change in edits.items():
            idx = int(idx_str)
            for col, val in change.items():
                st.session_state['df_ebay_ko'].at[idx, col] = val

def sync_ebay_en_editor():
    if "ebay_en_editor" in st.session_state:
        edits = st.session_state["ebay_en_editor"]["edited_rows"]
        for idx_str, change in edits.items():
            idx = int(idx_str)
            for col, val in change.items():
                st.session_state['df_ebay_en'].at[idx, col] = val

col_h1, col_h2 = st.columns([10, 2])
with col_h1:
    st.markdown('<h1 class="main-header" style="text-align: left; border-bottom: none;">Auction Monitor</h1>', unsafe_allow_html=True)
    st.markdown('<p class="sub-header" style="text-align: left;">진행 중이거나 진행 예정인 국내 경매와 이베이 자료를 찾는 페이지 입니다. <실시간 데이터 수집>버튼을 눌러보세요</p>', unsafe_allow_html=True)
with col_h2:
    if st.button("🏠 홈 (초기화)", use_container_width=True):
        st.session_state.clear()
        st.rerun()

default_ebay_ko = '빈티지, 의약, 약국, 약, 약제상'
default_ebay_en = "apothecary, bottle, jar, tool, scale, medicine"

col_kw1, col_kw2 = st.columns(2)
with col_kw1: ebay_keyword_ko = st.text_input("🔍 이베이 한국어 검색어", value=default_ebay_ko)
with col_kw2: ebay_keyword_en = st.text_input("🔍 이베이 영어 검색어", value=default_ebay_en)

col_sc1, col_sc2, col_sc3 = st.columns([5, 2, 5])
with col_sc2:
    if st.button("🚀 실시간 데이터 수집", use_container_width=True):
        st.session_state['ebay_ko_excel_ready'] = False
        st.session_state['ebay_en_excel_ready'] = False
        
        with st.status("데이터 수집 중...", expanded=True) as status:
            st.write("🏃 서울옥션 수집 중...")
            st.session_state['df_seoul'] = asyncio.run(async_scrape_seoul())
            st.write("🏃 칸옥션/마이아트 수집 중...")
            st.session_state['df_kan'] = asyncio.run(async_scrape_kan())
            st.session_state['df_myart'] = asyncio.run(async_scrape_myart())
            st.write("🏃 이베이 수집 중...")
            st.session_state['df_ebay_ko'] = scrape_ebay(ebay_keyword_ko)
            st.session_state['df_ebay_en'] = scrape_ebay(ebay_keyword_en)
            status.update(label="✅ 수집 완료!", state="complete", expanded=False)

if "df_seoul" in st.session_state:
    st.divider()
    
    col_dl1, col_dl2 = st.columns([10, 2])
    with col_dl2:
        if st.button("📊 통합 엑셀 생성", use_container_width=True):
            dfs = {"서울옥션": st.session_state.get('df_seoul'), "칸옥션": st.session_state.get('df_kan'), "마이아트옥션": st.session_state.get('df_myart'), "이베이(한국어)": st.session_state.get('df_ebay_ko'), "이베이(영어)": st.session_state.get('df_ebay_en')}
            all_excel_bytes = get_multi_sheet_excel(dfs)
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                tmp.write(all_excel_bytes)
                st.session_state['all_excel_path'] = tmp.name
            st.session_state['all_excel_ready'] = True
        
        if st.session_state.get('all_excel_ready'):
            with open(st.session_state['all_excel_path'], "rb") as f:
                st.download_button("📥 통합 엑셀 다운로드", data=f.read(), file_name="all_auctions.xlsx", use_container_width=True)
                
    st.divider()

    tabs = st.tabs(["🏛️ 서울옥션", "🏛️ 칸옥션", "🏛️ 마이아트옥션", "🛒 이베이(한국어)", "🛒 이베이(영어)"])
    
    with tabs[0]:
        st.data_editor(st.session_state.get('df_seoul', pd.DataFrame()), use_container_width=True, hide_index=True, height=600)

    
    with tabs[1]:
        st.data_editor(st.session_state.get('df_kan', pd.DataFrame()), column_config={"바로가기 URL": st.column_config.LinkColumn("상세")}, use_container_width=True, hide_index=True, height=600)

    with tabs[2]:
        st.data_editor(st.session_state.get('df_myart', pd.DataFrame()), column_config={"바로가기 URL": st.column_config.LinkColumn("상세")}, use_container_width=True, hide_index=True, height=600)

    with tabs[3]:
        col_s, col_r, _ = st.columns([1.5, 1.5, 7])
        with col_s: chk_all_ko = st.checkbox("전체 선택", key="chk_all_ko")
        if chk_all_ko != st.session_state.get('prev_chk_ko', False):
            st.session_state['df_ebay_ko']['선택'] = chk_all_ko
            st.session_state['prev_chk_ko'] = chk_all_ko
        with col_r:
            if st.button("✨ 연관성 높음", key="btn_rel_ko"):
                rel_kws = ["약장", "약합", "약재", "십장생", "건강", "장수", "의원", "의서", "약학", "의학", "유의", "약"]
                st.session_state['df_ebay_ko']['선택'] = st.session_state['df_ebay_ko'].apply(lambda x: any(kw in f"{x['항목명']}" for kw in rel_kws), axis=1)
                st.toast("✅ 연관 항목 선택 완료")
        st.data_editor(st.session_state['df_ebay_ko'], column_config={"바로가기": st.column_config.LinkColumn("이동"), "이미지": st.column_config.ImageColumn("이미지")}, use_container_width=True, hide_index=True, height=600, key="ebay_ko_editor", on_change=sync_ebay_ko_editor)
        
        selected_ko = st.session_state['df_ebay_ko'][st.session_state['df_ebay_ko']['선택'] == True]
        if not selected_ko.empty:
            st.markdown("<br>", unsafe_allow_html=True)
            ko_cols1, ko_cols2 = st.columns([10, 2])
            with ko_cols2:
                excel_bytes_ko = get_excel_data(selected_ko, "이베이(한국어)")
                st.download_button("📥 선택 항목 엑셀 다운로드", data=excel_bytes_ko, file_name="ebay_ko_selected.xlsx", key="dl_ebay_ko", use_container_width=True)

    with tabs[4]:
        col_s, _ = st.columns([1.5, 8.5])
        with col_s: chk_all_en = st.checkbox("전체 선택", key="chk_all_en")
        if chk_all_en != st.session_state.get('prev_chk_en', False):
            st.session_state['df_ebay_en']['선택'] = chk_all_en
            st.session_state['prev_chk_en'] = chk_all_en
        st.data_editor(st.session_state['df_ebay_en'], column_config={"바로가기": st.column_config.LinkColumn("이동"), "이미지": st.column_config.ImageColumn("이미지")}, use_container_width=True, hide_index=True, height=600, key="ebay_en_editor", on_change=sync_ebay_en_editor)

        selected_en = st.session_state['df_ebay_en'][st.session_state['df_ebay_en']['선택'] == True]
        if not selected_en.empty:
            st.markdown("<br>", unsafe_allow_html=True)
            en_cols1, en_cols2 = st.columns([10, 2])
            with en_cols2:
                excel_bytes_en = get_excel_data(selected_en, "이베이(영어)")
                st.download_button("📥 선택 항목 엑셀 다운로드", data=excel_bytes_en, file_name="ebay_en_selected.xlsx", key="dl_ebay_en", use_container_width=True)
